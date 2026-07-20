from datetime import date
from io import BytesIO

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.urls import reverse
from openpyxl import Workbook, load_workbook

from apps.backups.models import BackupJob, BackupJobTarget, BackupTechnology
from apps.customers.models import ManagedCustomer, Site
from apps.imports.models import ImportBatch, ImportRow
from apps.imports.services import (
    confirm_import_batch,
    create_import_preview,
    mark_import_batch_rolled_back,
)
from apps.inventory.models import ProtectedObject
from apps.operations.models import DailyControlEntry
from apps.tenancy.models import Membership, Organization


class PhaseThreeImportAndManualControlTest(TestCase):
    def setUp(self):
        User = get_user_model()
        self.user = User.objects.create_user(username="admin", password="pass")
        self.other_user = User.objects.create_user(username="other", password="pass")
        self.org = Organization.objects.create(name="Org", slug="org")
        self.other_org = Organization.objects.create(name="Other", slug="other")
        Membership.objects.create(
            organization=self.org,
            user=self.user,
            role=Membership.Role.ADMIN,
        )
        Membership.objects.create(
            organization=self.other_org,
            user=self.other_user,
            role=Membership.Role.ADMIN,
        )
        self.customer = ManagedCustomer.objects.create(organization=self.org, name="Cliente")
        self.site = Site.objects.create(
            organization=self.org,
            managed_customer=self.customer,
            name="MDP",
        )
        self.technology = BackupTechnology.objects.create(organization=self.org, name="Veeam")
        self.job = BackupJob.objects.create(
            organization=self.org,
            managed_customer=self.customer,
            site=self.site,
            technology=self.technology,
            name="Veeam diario",
        )
        self.server = ProtectedObject.objects.create(
            organization=self.org,
            managed_customer=self.customer,
            site=self.site,
            name="YAPP",
            external_reference="YA01V",
            object_type=ProtectedObject.ObjectType.PHYSICAL_SERVER,
        )
        self.target = BackupJobTarget.objects.create(
            organization=self.org,
            backup_job=self.job,
            protected_object=self.server,
        )
        self.other_customer = ManagedCustomer.objects.create(
            organization=self.other_org,
            name="Cliente ajeno",
        )
        self.other_site = Site.objects.create(
            organization=self.other_org,
            managed_customer=self.other_customer,
            name="BA",
        )
        self.other_technology = BackupTechnology.objects.create(
            organization=self.other_org,
            name="Iperius",
        )
        self.other_job = BackupJob.objects.create(
            organization=self.other_org,
            managed_customer=self.other_customer,
            site=self.other_site,
            technology=self.other_technology,
            name="Oculta",
        )
        self.client.force_login(self.user)

    def _workbook(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Cliente", "Sede", "Referencia", "Objeto", "Tecnología", "Tarea"])
        sheet.append(["Canteras", "MDP", "YA01V", "YAPP", "Veeam", "Veeam diario"])
        sheet.append([None, "MDP", "YA01V", "SQL Server", "Veeam", "SQL diario"])
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return output

    def _mapping(self):
        return {
            "customer": "Cliente",
            "site": "Sede",
            "reference": "Referencia",
            "object_name": "Objeto",
            "object_type": "",
            "technology": "Tecnología",
            "job_name": "Tarea",
            "status": "",
            "observation": "",
        }

    def test_import_preview_forward_fills_customer_cells(self):
        result = create_import_preview(
            organization=self.org,
            user=self.user,
            uploaded_file=self._workbook(),
            original_filename="control.xlsx",
            column_mapping=self._mapping(),
        )
        assert result.batch.row_count == 2
        rows = list(result.batch.rows.all())
        assert rows[0].normalized_data["customer"] == "Canteras"
        assert rows[1].normalized_data["customer"] == "Canteras"
        assert rows[1].normalized_data["object_name"] == "SQL Server"

    def test_confirm_import_creates_domain_records_and_targets(self):
        result = create_import_preview(
            organization=self.org,
            user=self.user,
            uploaded_file=self._workbook(),
            original_filename="control.xlsx",
            column_mapping=self._mapping(),
        )
        batch = confirm_import_batch(batch=result.batch, user=self.user)
        assert batch.status == ImportBatch.Status.IMPORTED
        assert ManagedCustomer.objects.filter(organization=self.org, name="Canteras").exists()
        assert ProtectedObject.objects.filter(
            organization=self.org,
            name="SQL Server",
            external_reference="YA01V",
        ).exists()
        assert BackupJob.objects.filter(organization=self.org, name="SQL diario").exists()
        assert BackupJobTarget.objects.filter(
            organization=self.org,
            backup_job__name="SQL diario",
            protected_object__name="SQL Server",
        ).exists()
        assert list(batch.rows.values_list("status", flat=True)) == [
            ImportRow.Status.IMPORTED,
            ImportRow.Status.IMPORTED,
        ]

    def test_import_batch_can_record_rollback_metadata(self):
        result = create_import_preview(
            organization=self.org,
            user=self.user,
            uploaded_file=self._workbook(),
            original_filename="control.xlsx",
            column_mapping=self._mapping(),
        )
        batch = confirm_import_batch(batch=result.batch, user=self.user)
        rolled_back = mark_import_batch_rolled_back(batch=batch, user=self.user)
        assert rolled_back.status == ImportBatch.Status.ROLLED_BACK
        assert rolled_back.rolled_back_by == self.user
        assert "Rollback metadata recorded." in rolled_back.notes

    def test_repeated_external_reference_is_allowed(self):
        ProtectedObject.objects.create(
            organization=self.org,
            managed_customer=self.customer,
            site=self.site,
            name="SQL Server",
            external_reference="YA01V",
            object_type=ProtectedObject.ObjectType.DATABASE,
        )
        assert ProtectedObject.objects.filter(
            organization=self.org,
            external_reference="YA01V",
        ).count() == 2

    def test_daily_control_entry_uses_target_object_and_org(self):
        entry = DailyControlEntry.objects.create(
            organization=self.org,
            control_date=date(2026, 7, 20),
            backup_job=self.job,
            backup_job_target=self.target,
            result=DailyControlEntry.Result.SUCCESS,
            manual_observation="Correcto manual",
            operator=self.user,
        )
        assert entry.organization == self.org
        assert entry.protected_object == self.server

    def test_daily_control_list_hides_other_organization_records(self):
        DailyControlEntry.objects.create(
            organization=self.org,
            control_date=date(2026, 7, 20),
            backup_job=self.job,
            backup_job_target=self.target,
            result=DailyControlEntry.Result.SUCCESS,
        )
        DailyControlEntry.objects.create(
            organization=self.other_org,
            control_date=date(2026, 7, 20),
            backup_job=self.other_job,
            result=DailyControlEntry.Result.ERROR,
            manual_observation="No visible",
        )
        response = self.client.get(reverse("daily_control_list"), {"date": "2026-07-20"})
        assert response.status_code == 200
        content = response.content.decode()
        assert "Veeam diario" in content
        assert "Oculta" not in content
        assert "No visible" not in content

    def test_daily_control_create_view_records_manual_result(self):
        response = self.client.post(
            reverse("daily_control_create"),
            {
                "control_date": "2026-07-20",
                "backup_job": str(self.job.id),
                "backup_job_target": str(self.target.id),
                "result": DailyControlEntry.Result.WARNING,
                "observed_at": "",
                "manual_observation": "Warning revisado",
                "ticket_reference": "",
            },
        )
        assert response.status_code == 302
        assert DailyControlEntry.objects.filter(
            organization=self.org,
            backup_job=self.job,
            protected_object=self.server,
            result=DailyControlEntry.Result.WARNING,
        ).exists()

    def test_daily_control_export_contains_only_active_organization_rows(self):
        DailyControlEntry.objects.create(
            organization=self.org,
            control_date=date(2026, 7, 20),
            backup_job=self.job,
            backup_job_target=self.target,
            result=DailyControlEntry.Result.SUCCESS,
            manual_observation="Visible",
        )
        DailyControlEntry.objects.create(
            organization=self.other_org,
            control_date=date(2026, 7, 20),
            backup_job=self.other_job,
            result=DailyControlEntry.Result.ERROR,
            manual_observation="Hidden",
        )
        response = self.client.get(reverse("daily_control_export"), {"date": "2026-07-20"})
        assert response.status_code == 200
        workbook = load_workbook(BytesIO(response.content))
        rows = list(workbook.active.iter_rows(values_only=True))
        assert rows[0][0] == "Fecha"
        assert any("Veeam diario" in row for row in rows[1:])
        assert not any("Oculta" in row for row in rows[1:])
