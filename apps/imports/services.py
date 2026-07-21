"""Spreadsheet import preview and confirmation services."""
from __future__ import annotations

from dataclasses import dataclass
from typing import IO, Any

from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from apps.backups.models import BackupJob, BackupJobTarget, BackupSchedule, BackupTechnology
from apps.customers.models import ManagedCustomer, Site
from apps.imports.models import ImportBatch, ImportRow, LegacyBackupConfiguration
from apps.inventory.models import ProtectedObject
from apps.tenancy.models import Organization


REQUIRED_FIELDS = ("customer", "site", "object_name", "job_name", "technology")
OPTIONAL_FIELDS = ("reference", "object_type", "status", "observation")


@dataclass(frozen=True)
class ImportPreviewResult:
    batch: ImportBatch
    rows: list[ImportRow]


def _clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _header_map(headers: list[str]) -> dict[str, int]:
    return {header: index for index, header in enumerate(headers) if header}


def _cell(row: tuple[Any, ...], headers: dict[str, int], column_name: str) -> str:
    index = headers.get(column_name)
    if index is None or index >= len(row):
        return ""
    return _clean(row[index])


def _status_for_row(organization: Organization, normalized: dict[str, str]) -> tuple[str, list[str]]:
    messages: list[str] = []
    missing = [field for field in REQUIRED_FIELDS if not normalized.get(field)]
    if missing:
        messages.append(f"Missing required fields: {', '.join(missing)}")
        return ImportRow.Status.INCOMPLETE, messages

    customer = ManagedCustomer.objects.filter(
        organization=organization,
        name=normalized["customer"],
    ).first()
    if not customer:
        messages.append("Managed customer will be created.")
        return ImportRow.Status.NEW, messages

    site = Site.objects.filter(
        organization=organization,
        managed_customer=customer,
        name=normalized["site"],
    ).first()
    if not site:
        messages.append("Site will be created.")
        return ImportRow.Status.NEW, messages

    protected_object = ProtectedObject.objects.filter(
        organization=organization,
        managed_customer=customer,
        site=site,
        name=normalized["object_name"],
    ).first()
    technology = BackupTechnology.objects.filter(
        organization=organization,
        name=normalized["technology"],
    ).first()
    job = None
    if technology:
        job = BackupJob.objects.filter(
            organization=organization,
            managed_customer=customer,
            site=site,
            technology=technology,
            name=normalized["job_name"],
        ).first()

    if protected_object and technology and job:
        messages.append("Existing customer, site, object, technology, and job found.")
        return ImportRow.Status.MATCHED, messages

    messages.append("One or more domain records will be created.")
    return ImportRow.Status.NEW, messages


@transaction.atomic
def create_import_preview(
    *,
    organization: Organization,
    user,
    uploaded_file: IO[bytes],
    original_filename: str,
    column_mapping: dict[str, str],
) -> ImportPreviewResult:
    workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    headers = [_clean(value) for value in next(rows_iter, ())]
    header_indexes = _header_map(headers)

    batch = ImportBatch.objects.create(
        organization=organization,
        original_filename=original_filename,
        column_mapping=column_mapping,
        created_by=user if getattr(user, "is_authenticated", False) else None,
    )

    preview_rows: list[ImportRow] = []
    last_customer = ""
    for row_number, row in enumerate(rows_iter, start=2):
        raw_data = {header: _clean(row[index]) if index < len(row) else "" for header, index in header_indexes.items()}
        normalized: dict[str, str] = {}
        for field in (*REQUIRED_FIELDS, *OPTIONAL_FIELDS):
            mapped_column = column_mapping.get(field, "")
            normalized[field] = _cell(row, header_indexes, mapped_column) if mapped_column else ""

        if normalized.get("customer"):
            last_customer = normalized["customer"]
        else:
            normalized["customer"] = last_customer

        if not normalized.get("object_type"):
            normalized["object_type"] = ProtectedObject.ObjectType.OTHER

        status, messages = _status_for_row(organization, normalized)
        preview_rows.append(
            ImportRow(
                organization=organization,
                batch=batch,
                row_number=row_number,
                raw_data=raw_data,
                normalized_data=normalized,
                status=status,
                messages=messages,
            )
        )

    ImportRow.objects.bulk_create(preview_rows)
    batch.row_count = len(preview_rows)
    batch.save(update_fields=["row_count", "updated_at"])
    return ImportPreviewResult(batch=batch, rows=list(batch.rows.all()))


@transaction.atomic
def confirm_import_batch(*, batch: ImportBatch, user) -> ImportBatch:
    if batch.status != ImportBatch.Status.PREVIEW:
        raise ValueError("Only preview batches can be confirmed.")

    imported_customer_ids: list[str] = []
    imported_site_ids: list[str] = []
    imported_object_ids: list[str] = []
    imported_technology_ids: list[str] = []
    imported_job_ids: list[str] = []
    imported_target_ids: list[str] = []

    for row in batch.rows.select_for_update().all():
        if row.status == ImportRow.Status.INCOMPLETE:
            row.status = ImportRow.Status.SKIPPED
            row.messages = [*row.messages, "Skipped because required fields are missing."]
            row.save(update_fields=["status", "messages", "updated_at"])
            continue

        data = row.normalized_data
        customer, customer_created = ManagedCustomer.objects.get_or_create(
            organization=batch.organization,
            name=data["customer"],
        )
        if customer_created:
            imported_customer_ids.append(str(customer.id))

        site, site_created = Site.objects.get_or_create(
            organization=batch.organization,
            managed_customer=customer,
            name=data["site"],
        )
        if site_created:
            imported_site_ids.append(str(site.id))

        technology, technology_created = BackupTechnology.objects.get_or_create(
            organization=batch.organization,
            name=data["technology"],
        )
        if technology_created:
            imported_technology_ids.append(str(technology.id))

        protected_object, object_created = ProtectedObject.objects.get_or_create(
            organization=batch.organization,
            managed_customer=customer,
            site=site,
            name=data["object_name"],
            defaults={
                "external_reference": data.get("reference", ""),
                "object_type": data.get("object_type") or ProtectedObject.ObjectType.OTHER,
            },
        )
        if object_created:
            imported_object_ids.append(str(protected_object.id))

        job, job_created = BackupJob.objects.get_or_create(
            organization=batch.organization,
            managed_customer=customer,
            site=site,
            technology=technology,
            name=data["job_name"],
        )
        if job_created:
            imported_job_ids.append(str(job.id))

        target, target_created = BackupJobTarget.objects.get_or_create(
            organization=batch.organization,
            backup_job=job,
            protected_object=protected_object,
        )
        if target_created:
            imported_target_ids.append(str(target.id))

        row.status = ImportRow.Status.IMPORTED
        row.messages = [*row.messages, "Imported or matched successfully."]
        row.save(update_fields=["status", "messages", "updated_at"])

    batch.imported_customer_ids = imported_customer_ids
    batch.imported_site_ids = imported_site_ids
    batch.imported_object_ids = imported_object_ids
    batch.imported_technology_ids = imported_technology_ids
    batch.imported_job_ids = imported_job_ids
    batch.imported_target_ids = imported_target_ids
    batch.status = ImportBatch.Status.IMPORTED
    batch.confirmed_by = user if getattr(user, "is_authenticated", False) else None
    batch.confirmed_at = timezone.now()
    batch.save()
    return batch


@transaction.atomic
def mark_import_batch_rolled_back(*, batch: ImportBatch, user) -> ImportBatch:
    if batch.status != ImportBatch.Status.IMPORTED:
        raise ValueError("Only imported batches can be marked as rolled back.")
    batch.status = ImportBatch.Status.ROLLED_BACK
    batch.rolled_back_by = user if getattr(user, "is_authenticated", False) else None
    batch.rolled_back_at = timezone.now()
    batch.notes = (batch.notes + "\n" if batch.notes else "") + "Rollback metadata recorded."
    batch.save()
    return batch


@dataclass(frozen=True)
class LegacyReconciliationResult:
    site: Site
    technology: BackupTechnology
    protected_object: ProtectedObject
    backup_job: BackupJob
    target: BackupJobTarget
    created: dict[str, bool]


@transaction.atomic
def reconcile_legacy_configuration(
    *,
    legacy_configuration,
    site_name: str,
    object_name: str,
    technology_name: str,
    job_name: str,
    user=None,
    note: str = "",
) -> LegacyReconciliationResult:
    site_name = _clean(site_name)
    object_name = _clean(object_name)
    technology_name = _clean(technology_name)
    job_name = _clean(job_name)
    if not all((site_name, object_name, technology_name, job_name)):
        raise ValueError("Site, object, technology, and job names are required.")

    organization = legacy_configuration.organization
    customer = legacy_configuration.managed_customer
    site, site_created = Site.objects.get_or_create(
        organization=organization,
        managed_customer=customer,
        name=site_name,
    )
    technology, technology_created = BackupTechnology.objects.get_or_create(
        organization=organization,
        name=technology_name,
    )
    protected_object, object_created = ProtectedObject.objects.get_or_create(
        organization=organization,
        managed_customer=customer,
        site=site,
        name=object_name,
        defaults={
            "object_type": ProtectedObject.ObjectType.OTHER,
            "notes": f"Reconciled from legacy CSV row {legacy_configuration.source_row}.",
        },
    )
    job, job_created = BackupJob.objects.get_or_create(
        organization=organization,
        managed_customer=customer,
        site=site,
        technology=technology,
        name=job_name,
        defaults={
            "matching_aliases": _legacy_aliases(legacy_configuration),
            "notes": f"Reconciled from legacy CSV row {legacy_configuration.source_row}.",
        },
    )
    target, target_created = BackupJobTarget.objects.get_or_create(
        organization=organization,
        backup_job=job,
        protected_object=protected_object,
    )

    legacy_configuration.reconciled_site = site
    legacy_configuration.reconciled_protected_object = protected_object
    legacy_configuration.reconciled_backup_job = job
    legacy_configuration.reconciled_by = user if getattr(user, "is_authenticated", False) else None
    legacy_configuration.reconciled_at = timezone.now()
    legacy_configuration.reconciliation_note = note
    legacy_configuration.save(
        update_fields=[
            "reconciled_site",
            "reconciled_protected_object",
            "reconciled_backup_job",
            "reconciled_by",
            "reconciled_at",
            "reconciliation_note",
            "updated_at",
        ]
    )
    return LegacyReconciliationResult(
        site=site,
        technology=technology,
        protected_object=protected_object,
        backup_job=job,
        target=target,
        created={
            "site": site_created,
            "technology": technology_created,
            "protected_object": object_created,
            "backup_job": job_created,
            "target": target_created,
        },
    )


def _legacy_aliases(legacy_configuration) -> str:
    aliases = [
        legacy_configuration.legacy_backup_name,
        legacy_configuration.source_asset_label,
    ]
    return "\n".join(alias for alias in aliases if alias)


@dataclass(frozen=True)
class LegacyBootstrapResult:
    processed: int
    skipped_already_reconciled: int
    sites_created: int
    technologies_created: int
    protected_objects_created: int
    backup_jobs_created: int
    targets_created: int

    def to_dict(self, *, tenant: str, source_sha256: str) -> dict[str, Any]:
        return {
            "tenant": tenant,
            "source_sha256": source_sha256,
            "processed": self.processed,
            "skipped_already_reconciled": self.skipped_already_reconciled,
            "sites_created": self.sites_created,
            "technologies_created": self.technologies_created,
            "protected_objects_created": self.protected_objects_created,
            "backup_jobs_created": self.backup_jobs_created,
            "targets_created": self.targets_created,
            "creates_schedules": False,
            "creates_expected_executions": False,
            "requires_review": True,
        }


@transaction.atomic
def bootstrap_legacy_backups(*, organization: Organization, source_sha256: str) -> LegacyBootstrapResult:
    configurations = LegacyBackupConfiguration.objects.select_for_update().filter(
        organization=organization,
        source_sha256=source_sha256,
    )
    counters = {
        "processed": 0,
        "skipped_already_reconciled": 0,
        "sites_created": 0,
        "technologies_created": 0,
        "protected_objects_created": 0,
        "backup_jobs_created": 0,
        "targets_created": 0,
    }
    for configuration in configurations:
        if configuration.reconciled_backup_job_id:
            counters["skipped_already_reconciled"] += 1
            continue
        result = reconcile_legacy_configuration(
            legacy_configuration=configuration,
            site_name=_bootstrap_site_name(configuration),
            object_name=_bootstrap_object_name(configuration),
            technology_name=_bootstrap_technology_name(configuration),
            job_name=_bootstrap_job_name(configuration),
            note="Provisional bootstrap from legacy import; requires review.",
        )
        counters["processed"] += 1
        for key, created in result.created.items():
            counter_key = {
                "site": "sites_created",
                "technology": "technologies_created",
                "protected_object": "protected_objects_created",
                "backup_job": "backup_jobs_created",
                "target": "targets_created",
            }[key]
            if created:
                counters[counter_key] += 1
    return LegacyBootstrapResult(**counters)


def _bootstrap_site_name(configuration: LegacyBackupConfiguration) -> str:
    site = _clean(configuration.legacy_site_label)
    if site.casefold() in {"", "n/a", "na"}:
        return "General"
    return site


def _bootstrap_object_name(configuration: LegacyBackupConfiguration) -> str:
    source_asset = _clean(configuration.source_asset_label)
    if source_asset:
        return source_asset
    backup_name = _clean(configuration.legacy_backup_name)
    if backup_name:
        return f"{backup_name} [legacy row {configuration.source_row}]"
    return f"Legacy source row {configuration.source_row}"


def _bootstrap_technology_name(configuration: LegacyBackupConfiguration) -> str:
    method = _clean(configuration.legacy_method)
    if method:
        return method
    provider = _clean(configuration.provider)
    if provider:
        return provider
    return "Legacy method pending review"


def _bootstrap_job_name(configuration: LegacyBackupConfiguration) -> str:
    base = _clean(configuration.legacy_backup_name) or _clean(configuration.source_asset_label)
    if not base:
        base = "Legacy backup"
    return f"{base} [legacy row {configuration.source_row}]"


@dataclass(frozen=True)
class LegacyScheduleBootstrapResult:
    considered: int
    created: int
    updated_existing: int
    skipped_existing: int

    def to_dict(self, *, tenant: str) -> dict[str, Any]:
        return {
            "tenant": tenant,
            "considered": self.considered,
            "created": self.created,
            "updated_existing": self.updated_existing,
            "skipped_existing": self.skipped_existing,
            "creates_expected_executions": False,
            "requires_review": True,
        }


@transaction.atomic
def bootstrap_legacy_schedules(
    *,
    organization: Organization,
    frequency: str,
    weekdays: str,
    scheduled_time,
    report_deadline_time,
    report_deadline_offset_days: int,
    mode: str,
    timezone_name: str,
    update_existing: bool = False,
) -> LegacyScheduleBootstrapResult:
    job_ids = list(
        LegacyBackupConfiguration.objects.filter(
            organization=organization,
            reconciled_backup_job__isnull=False,
        )
        .values_list("reconciled_backup_job_id", flat=True)
        .distinct()
    )
    jobs = BackupJob.objects.select_for_update().filter(
        organization=organization,
        id__in=job_ids,
    )
    considered = 0
    created = 0
    updated_existing = 0
    skipped_existing = 0
    for job in jobs:
        considered += 1
        existing_schedule = BackupSchedule.objects.filter(
            organization=organization,
            backup_job=job,
        ).first()
        if existing_schedule:
            if not update_existing:
                skipped_existing += 1
                continue
            existing_schedule.frequency = frequency
            existing_schedule.weekdays = weekdays
            existing_schedule.scheduled_time = scheduled_time
            existing_schedule.timezone = timezone_name
            existing_schedule.report_deadline_time = report_deadline_time
            existing_schedule.report_deadline_offset_days = report_deadline_offset_days
            existing_schedule.mode = mode
            if "Updated provisional schedule" not in existing_schedule.notes:
                existing_schedule.notes = (
                    (existing_schedule.notes + "\n") if existing_schedule.notes else ""
                ) + "Updated provisional schedule from legacy bootstrap; requires review."
            existing_schedule.save(
                update_fields=[
                    "frequency",
                    "weekdays",
                    "scheduled_time",
                    "timezone",
                    "report_deadline_time",
                    "report_deadline_offset_days",
                    "mode",
                    "notes",
                    "updated_at",
                ]
            )
            updated_existing += 1
            continue
        BackupSchedule.objects.create(
            organization=organization,
            backup_job=job,
            frequency=frequency,
            weekdays=weekdays,
            scheduled_time=scheduled_time,
            timezone=timezone_name,
            report_deadline_time=report_deadline_time,
            report_deadline_offset_days=report_deadline_offset_days,
            mode=mode,
            notes="Provisional schedule from legacy bootstrap; requires review.",
        )
        created += 1
    return LegacyScheduleBootstrapResult(
        considered=considered,
        created=created,
        updated_existing=updated_existing,
        skipped_existing=skipped_existing,
    )
